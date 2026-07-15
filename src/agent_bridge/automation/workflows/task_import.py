from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

MAX_IMPORT_ROWS = 5000
SUPPORTED_IMPORT_EXTENSION = ".xlsx"

_FIXED_HEADERS = ("task_key", "task_version", "type")


class TaskImportFormatError(ValueError):
    """工作簿无法生成有效的任务导入预览。"""


@dataclass(frozen=True)
class ParsedTaskImportRow:
    row_number: int
    task_key: str
    task_version: str
    task_type: str
    payload: dict[str, Any]
    errors: tuple[str, ...]


@dataclass(frozen=True)
class ParsedTaskImport:
    filename: str
    sheet_name: str
    rows: tuple[ParsedTaskImportRow, ...]


def parse_task_import(content: bytes, *, filename: str) -> ParsedTaskImport:
    """解析任务导入工作簿的第一个工作表。"""
    if not filename.lower().endswith(SUPPORTED_IMPORT_EXTENSION):
        raise TaskImportFormatError("仅支持 .xlsx 文件")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError, EOFError, KeyError) as exc:
        raise TaskImportFormatError("xlsx 文件无效") from exc

    try:
        if not workbook.worksheets:
            raise TaskImportFormatError("xlsx 文件没有工作表")
        worksheet = workbook.worksheets[0]
        row_iterator = worksheet.iter_rows()
        try:
            header_row = tuple(_read_cell(cell) for cell in next(row_iterator))
        except StopIteration as exc:
            raise TaskImportFormatError("缺少表头") from exc

        headers = _prepare_headers(header_row)
        fixed_indexes = {
            normalized: index
            for index, normalized, _display in headers
            if normalized in _FIXED_HEADERS
        }
        if "task_key" not in fixed_indexes:
            raise TaskImportFormatError("缺少 task_key 表头")

        rows: list[ParsedTaskImportRow] = []
        seen_keys: set[tuple[str, str]] = set()
        data_row_count = 0
        for row_number, cells in enumerate(row_iterator, start=2):
            values = tuple(_read_cell(cell) for cell in cells)
            if _row_is_empty(values):
                continue
            data_row_count += 1
            if data_row_count > MAX_IMPORT_ROWS:
                raise TaskImportFormatError(
                    f"数据行不能超过 {MAX_IMPORT_ROWS} 行（第{row_number}行）"
                )

            task_key = _text_value(_cell_at(values, fixed_indexes["task_key"]))
            task_version = _text_value(
                _cell_at(values, fixed_indexes.get("task_version"))
            )
            task_type = _text_value(_cell_at(values, fixed_indexes.get("type")))
            payload: dict[str, Any] = {}
            errors: list[str] = []

            if not task_key:
                errors.append("task_key 不能为空")

            for index, normalized, display in headers:
                value = _cell_at(values, index)
                if normalized in _FIXED_HEADERS:
                    continue
                if not normalized:
                    if not _is_empty_cell(value):
                        errors.append(f"第{index + 1}列缺少表头")
                    continue
                if _is_empty_cell(value):
                    continue
                payload[display] = _json_scalar(value)

            if task_key:
                key = (task_key, task_version)
                if key in seen_keys:
                    errors.append("task_key + task_version 重复")
                else:
                    seen_keys.add(key)

            rows.append(
                ParsedTaskImportRow(
                    row_number=row_number,
                    task_key=task_key,
                    task_version=task_version,
                    task_type=task_type,
                    payload=payload,
                    errors=tuple(errors),
                )
            )

        return ParsedTaskImport(
            filename=filename,
            sheet_name=worksheet.title,
            rows=tuple(rows),
        )
    finally:
        workbook.close()


def build_task_import_template() -> bytes:
    """生成包含固定列和说明页的任务导入模板。"""
    workbook = Workbook()
    tasks = workbook.active
    tasks.title = "tasks"
    tasks.append(list(_FIXED_HEADERS))

    explanation = workbook.create_sheet("说明")
    explanation.append(["任务导入说明"])
    explanation.append(["固定列：task_key、task_version、type。"])
    explanation.append(["额外列的列名会进入 payload，单元格值会作为对应的 payload 值。"])

    output = BytesIO()
    try:
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()


def _prepare_headers(header_row: tuple[Any, ...]) -> list[tuple[int, str, str]]:
    headers: list[tuple[int, str, str]] = []
    seen: dict[str, str] = {}
    for index, value in enumerate(header_row):
        normalized = _normalize_header(value)
        display = _display_header(value)
        if normalized and normalized in seen:
            raise TaskImportFormatError(f"表头重复: {display or seen[normalized]}")
        if normalized:
            seen[normalized] = display
        headers.append((index, normalized, display))
    return headers


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lstrip("\ufeff").strip().casefold()


def _display_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lstrip("\ufeff").strip()


def _cell_at(values: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(values):
        return None
    return values[index]


def _read_cell(cell: Any) -> Any:
    value = cell.value
    if (
        isinstance(value, datetime)
        and getattr(cell, "is_date", False)
        and not _number_format_has_time(getattr(cell, "number_format", ""))
    ):
        return value.date()
    return value


def _number_format_has_time(number_format: str) -> bool:
    normalized = number_format.lower().replace(" ", "")
    return "h" in normalized or "s" in normalized or "am/pm" in normalized


def _row_is_empty(values: tuple[Any, ...]) -> bool:
    return all(_is_empty_cell(value) for value in values)


def _is_empty_cell(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _text_value(value: Any) -> str:
    scalar = _json_scalar(value)
    return "" if scalar is None else str(scalar).strip()


def _json_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
