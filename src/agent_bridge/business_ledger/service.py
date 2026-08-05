"""业务台账的定义、持久化与内存快照查询。"""

from __future__ import annotations

import asyncio
from io import BytesIO
import json
import sqlite3
import threading
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook

from agent_bridge.core.domain import NotFound, ValidationError, require_admin_user
from agent_bridge.core.editing import attach_edit_token, require_edit_token


MAX_LEDGER_FIELDS = 100
MAX_LEDGER_RECORDS = 50_000
FIELD_TYPES = {"text", "number", "enum", "date", "datetime"}
TEXT_OPS = {"exact", "prefix", "contains"}
NUMBER_OPS = {"exact", "gt", "gte", "lt", "lte", "between"}
DATE_OPS = {"exact", "before", "after", "between"}
ENUM_OPS = {"exact", "in"}


LEDGER_SCHEMA = """
CREATE TABLE IF NOT EXISTS business_ledgers (
  ledger_key TEXT PRIMARY KEY,
  name TEXT NOT NULL,
  description TEXT NOT NULL DEFAULT '',
  definition_json TEXT NOT NULL,
  current_revision_no INTEGER NOT NULL DEFAULT 1,
  created_by TEXT NOT NULL,
  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
CREATE TABLE IF NOT EXISTS business_ledger_revisions (
  ledger_key TEXT NOT NULL REFERENCES business_ledgers(ledger_key) ON DELETE CASCADE,
  revision_no INTEGER NOT NULL,
  definition_json TEXT NOT NULL,
  created_by TEXT NOT NULL,
  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
  PRIMARY KEY (ledger_key, revision_no)
);
CREATE TABLE IF NOT EXISTS business_ledger_records (
  ledger_key TEXT NOT NULL REFERENCES business_ledgers(ledger_key) ON DELETE CASCADE,
  record_id TEXT NOT NULL,
  values_json TEXT NOT NULL,
  created_by TEXT NOT NULL,
  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
  updated_by TEXT NOT NULL,
  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
  PRIMARY KEY (ledger_key, record_id)
);
CREATE INDEX IF NOT EXISTS idx_business_ledger_records_ledger ON business_ledger_records(ledger_key, updated_at DESC);
"""


@dataclass(frozen=True)
class LedgerSnapshot:
    definition: dict[str, Any]
    frame: pd.DataFrame


class BusinessLedgerService:
    """管理台账数据；所有读取查询均针对不可变的内存快照。"""

    def __init__(self, *, db_path: Path, admins: set[str]) -> None:
        self.db_path = db_path
        self.admins = admins
        self._snapshots: dict[str, LedgerSnapshot] = {}
        self._import_previews: dict[str, tuple[str, list[dict[str, Any]]]] = {}
        self._lock = threading.RLock()

    def _connect(self):
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode = WAL")
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def init_schema(self) -> None:
        with self._connect() as conn:
            conn.executescript(LEDGER_SCHEMA)

    async def load_all_async(self) -> None:
        await asyncio.to_thread(self.reload_all)

    def reload_all(self) -> None:
        definitions = self._list_definitions_raw()
        snapshots = {item["ledger_key"]: self._build_snapshot(item) for item in definitions}
        with self._lock:
            self._snapshots = snapshots

    def list_ledgers(self, actor: str) -> list[dict[str, Any]]:
        require_admin_user(actor, self.admins)
        definitions = self._list_definitions_raw()
        with self._connect() as conn:
            counts = {
                row["ledger_key"]: int(row["count"])
                for row in conn.execute(
                    "SELECT ledger_key, COUNT(*) AS count FROM business_ledger_records GROUP BY ledger_key"
                ).fetchall()
            }
        return [self._payload(item, record_count=counts.get(item["ledger_key"], 0)) for item in definitions]

    def ledger_contexts(self, ledger_keys: list[str]) -> list[dict[str, Any]]:
        """返回可安全注入 Profile 的台账说明，不要求管理权限。"""
        allowed = set(ledger_keys)
        return [
            {
                "ledger_key": item["ledger_key"],
                "name": item["name"],
                "description": item["description"],
                "fields": [
                    {
                        "field_key": field["field_key"],
                        "name": field["name"],
                        "field_type": field["field_type"],
                        "query_modes": field["query_modes"],
                    }
                    for field in item["fields"]
                    if field["query_modes"] or field["agent_readable"]
                ],
            }
            for item in self._list_definitions_raw()
            if item["ledger_key"] in allowed
        ]

    def ledger_keys(self) -> list[str]:
        return [item["ledger_key"] for item in self._list_definitions_raw()]

    def get_ledger(self, actor: str, ledger_key: str) -> dict[str, Any]:
        require_admin_user(actor, self.admins)
        ledger = self._get_definition_raw(ledger_key)
        return self._payload(ledger, record_count=self._record_count(ledger_key))

    def create_ledger(
        self, actor: str, *, ledger_key: str, name: str, description: str, fields: list[dict[str, Any]], expected_edit_token: str | None = None
    ) -> dict[str, Any]:
        require_admin_user(actor, self.admins)
        self._validate_ledger_key(ledger_key)
        existing = self._get_definition_raw(ledger_key, required=False)
        require_edit_token(expected_edit_token, existing, resource_type="业务台账", resource_key=ledger_key, actor=actor)
        if existing is not None:
            raise ValidationError("台账标识已存在")
        definition = self._definition(ledger_key, name, description, fields)
        with self._connect() as conn:
            conn.execute(
                "INSERT INTO business_ledgers (ledger_key, name, description, definition_json, created_by) VALUES (?, ?, ?, ?, ?)",
                (ledger_key, name, description, json.dumps(definition, ensure_ascii=False), actor),
            )
            conn.execute(
                "INSERT INTO business_ledger_revisions (ledger_key, revision_no, definition_json, created_by) VALUES (?, 1, ?, ?)",
                (ledger_key, json.dumps(definition, ensure_ascii=False), actor),
            )
        self._replace_snapshot(ledger_key)
        return self.get_ledger(actor, ledger_key)

    def update_ledger(
        self, actor: str, ledger_key: str, *, name: str, description: str, fields: list[dict[str, Any]], expected_edit_token: str | None = None
    ) -> dict[str, Any]:
        require_admin_user(actor, self.admins)
        current = self._get_definition_raw(ledger_key)
        require_edit_token(expected_edit_token, current, resource_type="业务台账", resource_key=ledger_key, actor=actor)
        definition = self._definition(ledger_key, name, description, fields)
        # 已有记录必须可被新定义读取，防止类型修改造成不可见脏数据。
        for row in self._records_raw(ledger_key):
            self._normalize_values(definition, json.loads(row["values_json"]))
        revision_no = int(current["current_revision_no"]) + 1
        with self._connect() as conn:
            conn.execute(
                "UPDATE business_ledgers SET name = ?, description = ?, definition_json = ?, current_revision_no = ?, updated_at = CURRENT_TIMESTAMP WHERE ledger_key = ?",
                (name, description, json.dumps(definition, ensure_ascii=False), revision_no, ledger_key),
            )
            conn.execute(
                "INSERT INTO business_ledger_revisions (ledger_key, revision_no, definition_json, created_by) VALUES (?, ?, ?, ?)",
                (ledger_key, revision_no, json.dumps(definition, ensure_ascii=False), actor),
            )
        self._replace_snapshot(ledger_key)
        return self.get_ledger(actor, ledger_key)

    def delete_ledger(self, actor: str, ledger_key: str) -> None:
        require_admin_user(actor, self.admins)
        self._get_definition_raw(ledger_key)
        with self._connect() as conn:
            conn.execute("DELETE FROM business_ledgers WHERE ledger_key = ?", (ledger_key,))
        with self._lock:
            self._snapshots.pop(ledger_key, None)

    def list_records(self, actor: str, ledger_key: str, **query: Any) -> dict[str, Any]:
        require_admin_user(actor, self.admins)
        return self.query(ledger_key, **query)

    def add_record(self, actor: str, ledger_key: str, values: dict[str, Any]) -> dict[str, Any]:
        require_admin_user(actor, self.admins)
        definition = self._get_definition_raw(ledger_key)
        normalized = self._normalize_values(definition, values)
        if self._record_count(ledger_key) >= MAX_LEDGER_RECORDS:
            raise ValidationError(f"单个台账最多保存 {MAX_LEDGER_RECORDS} 行数据")
        record_id = uuid.uuid4().hex
        with self._connect() as conn:
            conn.execute(
                "INSERT INTO business_ledger_records (ledger_key, record_id, values_json, created_by, updated_by) VALUES (?, ?, ?, ?, ?)",
                (ledger_key, record_id, json.dumps(normalized, ensure_ascii=False), actor, actor),
            )
        self._replace_snapshot(ledger_key)
        return {"record_id": record_id, "values": normalized}

    def update_record(self, actor: str, ledger_key: str, record_id: str, values: dict[str, Any]) -> dict[str, Any]:
        require_admin_user(actor, self.admins)
        definition = self._get_definition_raw(ledger_key)
        normalized = self._normalize_values(definition, values)
        with self._connect() as conn:
            changed = conn.execute(
                "UPDATE business_ledger_records SET values_json = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP WHERE ledger_key = ? AND record_id = ?",
                (json.dumps(normalized, ensure_ascii=False), actor, ledger_key, record_id),
            ).rowcount
        if not changed:
            raise NotFound("台账记录不存在")
        self._replace_snapshot(ledger_key)
        return {"record_id": record_id, "values": normalized}

    def delete_record(self, actor: str, ledger_key: str, record_id: str) -> None:
        require_admin_user(actor, self.admins)
        with self._connect() as conn:
            changed = conn.execute(
                "DELETE FROM business_ledger_records WHERE ledger_key = ? AND record_id = ?", (ledger_key, record_id)
            ).rowcount
        if not changed:
            raise NotFound("台账记录不存在")
        self._replace_snapshot(ledger_key)

    def preview_xlsx_import(self, actor: str, ledger_key: str, content: bytes) -> dict[str, Any]:
        require_admin_user(actor, self.admins)
        definition = self._get_definition_raw(ledger_key)
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        except Exception as exc:
            raise ValidationError("xlsx 文件无效") from exc
        if not rows:
            raise ValidationError("xlsx 文件为空")
        headers = [str(value or "").strip() for value in rows[0]]
        fields = {field["field_key"]: field for field in definition["fields"]}
        names = {field["name"]: field["field_key"] for field in definition["fields"]}
        mapped = [fields.get(header, fields.get(names.get(header, ""))) for header in headers]
        unknown = [header for header, field in zip(headers, mapped) if header and field is None]
        if unknown:
            raise ValidationError(f"Excel 存在未匹配列：{unknown[0]}")
        prepared: list[dict[str, Any]] = []
        errors: list[dict[str, Any]] = []
        for row_no, values in enumerate(rows[1:], start=2):
            if all(value is None for value in values):
                continue
            raw = {field["field_key"]: value for field, value in zip(mapped, values) if field is not None}
            try:
                prepared.append(self._normalize_values(definition, raw))
            except ValidationError as exc:
                errors.append({"row": row_no, "error": exc.message})
        preview_id = uuid.uuid4().hex
        self._import_previews[preview_id] = (ledger_key, prepared)
        return {"preview_id": preview_id, "rows": len(prepared), "errors": errors, "columns": [field["field_key"] if field else None for field in mapped]}

    def confirm_xlsx_import(self, actor: str, ledger_key: str, preview_id: str) -> dict[str, Any]:
        require_admin_user(actor, self.admins)
        prepared = self._import_previews.pop(preview_id, None)
        if prepared is None or prepared[0] != ledger_key:
            raise NotFound("导入预览不存在或已过期")
        rows = prepared[1]
        if self._record_count(ledger_key) + len(rows) > MAX_LEDGER_RECORDS:
            raise ValidationError(f"导入后超过单台账 {MAX_LEDGER_RECORDS} 行上限")
        with self._connect() as conn:
            conn.executemany(
                "INSERT INTO business_ledger_records (ledger_key, record_id, values_json, created_by, updated_by) VALUES (?, ?, ?, ?, ?)",
                [(ledger_key, uuid.uuid4().hex, json.dumps(values, ensure_ascii=False), actor, actor) for values in rows],
            )
        self._replace_snapshot(ledger_key)
        return {"ledger_key": ledger_key, "imported": len(rows)}

    def export_xlsx(self, actor: str, ledger_key: str) -> bytes:
        require_admin_user(actor, self.admins)
        snapshot = self._snapshot(ledger_key)
        workbook = Workbook()
        sheet = workbook.active
        fields = snapshot.definition["fields"]
        sheet.append([field["field_key"] for field in fields])
        for _, row in snapshot.frame.iterrows():
            sheet.append([self._json_value(row[field["field_key"]]) for field in fields])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def query(
        self,
        ledger_key: str,
        *,
        filters: dict[str, dict[str, Any]] | None = None,
        keyword: str | None = None,
        sort: dict[str, str] | None = None,
        limit: int = 50,
        offset: int = 0,
        agent_visible_only: bool = False,
    ) -> dict[str, Any]:
        snapshot = self._snapshot(ledger_key)
        definition = snapshot.definition
        fields = {field["field_key"]: field for field in definition["fields"]}
        frame = snapshot.frame
        mask = pd.Series(True, index=frame.index, dtype=bool)
        for field_key, condition in (filters or {}).items():
            field = fields.get(field_key)
            if field is None or not field["query_modes"]:
                raise ValidationError(f"字段不可查询：{field_key}")
            operator = str(condition.get("op") or "")
            if operator not in field["query_modes"]:
                raise ValidationError(f"字段不支持该查询方式：{field_key}/{operator}")
            mask &= self._condition_mask(frame[field_key], operator, condition)
        if keyword:
            searchable = [item["field_key"] for item in definition["fields"] if "contains" in item["query_modes"]]
            if not searchable:
                raise ValidationError("当前台账未配置关键词检索字段")
            keyword_mask = pd.Series(False, index=frame.index, dtype=bool)
            for field_key in searchable:
                keyword_mask |= frame[field_key].fillna("").astype(str).str.contains(str(keyword), regex=False, na=False)
            mask &= keyword_mask
        filtered = frame.loc[mask]
        if sort:
            sort_field = str(sort.get("field") or "")
            field = fields.get(sort_field)
            if field is None or not field.get("sortable"):
                raise ValidationError(f"字段不可排序：{sort_field}")
            filtered = filtered.sort_values(sort_field, ascending=str(sort.get("direction", "asc")).lower() != "desc", kind="stable")
        bounded_limit = min(max(int(limit), 1), 100)
        bounded_offset = max(int(offset), 0)
        visible_fields = [field["field_key"] for field in definition["fields"] if not agent_visible_only or field["agent_readable"]]
        rows = []
        for _, row in filtered.iloc[bounded_offset : bounded_offset + bounded_limit].iterrows():
            rows.append({"record_id": row["_record_id"], "values": {key: self._json_value(row[key]) for key in visible_fields}})
        return {"ledger_key": ledger_key, "total": int(len(filtered)), "limit": bounded_limit, "offset": bounded_offset, "items": rows}

    def _condition_mask(self, series: pd.Series, operator: str, condition: dict[str, Any]) -> pd.Series:
        value = condition.get("value")
        if operator == "exact":
            return series == value
        if operator == "in":
            values = condition.get("values", value)
            return series.isin(values if isinstance(values, list) else [values])
        if operator == "prefix":
            return series.fillna("").astype(str).str.startswith(str(value), na=False)
        if operator == "contains":
            return series.fillna("").astype(str).str.contains(str(value), regex=False, na=False)
        if operator in {"gt", "after"}:
            return series > value
        if operator in {"gte"}:
            return series >= value
        if operator in {"lt", "before"}:
            return series < value
        if operator in {"lte"}:
            return series <= value
        if operator == "between":
            lower = condition.get("from")
            upper = condition.get("to")
            return (series >= lower) & (series <= upper)
        raise ValidationError(f"不支持的查询方式：{operator}")

    def _snapshot(self, ledger_key: str) -> LedgerSnapshot:
        with self._lock:
            snapshot = self._snapshots.get(ledger_key)
        if snapshot is None:
            raise NotFound("业务台账不存在或仍在加载")
        return snapshot

    def _replace_snapshot(self, ledger_key: str) -> None:
        definition = self._get_definition_raw(ledger_key)
        snapshot = self._build_snapshot(definition)
        with self._lock:
            self._snapshots = {**self._snapshots, ledger_key: snapshot}

    def _build_snapshot(self, definition: dict[str, Any]) -> LedgerSnapshot:
        rows = []
        for record in self._records_raw(definition["ledger_key"]):
            values = json.loads(record["values_json"])
            rows.append({"_record_id": record["record_id"], **values})
        columns = ["_record_id", *[field["field_key"] for field in definition["fields"]]]
        frame = pd.DataFrame(rows, columns=columns)
        for field in definition["fields"]:
            key, field_type = field["field_key"], field["field_type"]
            if field_type == "number":
                frame[key] = pd.to_numeric(frame[key], errors="raise")
            elif field_type in {"date", "datetime"}:
                frame[key] = pd.to_datetime(frame[key], errors="raise", utc=field_type == "datetime")
        return LedgerSnapshot(definition=definition, frame=frame)

    def _definition(self, ledger_key: str, name: str, description: str, fields: list[dict[str, Any]]) -> dict[str, Any]:
        if not name.strip():
            raise ValidationError("台账名称不能为空")
        if not isinstance(fields, list) or not fields or len(fields) > MAX_LEDGER_FIELDS:
            raise ValidationError(f"台账字段数量必须为 1 到 {MAX_LEDGER_FIELDS}")
        normalized_fields = [self._normalize_field(field) for field in fields]
        if len({field["field_key"] for field in normalized_fields}) != len(normalized_fields):
            raise ValidationError("字段标识不可重复")
        return {"ledger_key": ledger_key, "name": name.strip(), "description": description.strip(), "fields": normalized_fields}

    def _normalize_field(self, field: dict[str, Any]) -> dict[str, Any]:
        key = str(field.get("field_key") or "").strip()
        self._validate_ledger_key(key, label="字段标识")
        field_type = str(field.get("field_type") or "").strip()
        if field_type not in FIELD_TYPES:
            raise ValidationError(f"不支持的字段类型：{field_type}")
        valid_ops = TEXT_OPS if field_type == "text" else NUMBER_OPS if field_type == "number" else ENUM_OPS if field_type == "enum" else DATE_OPS
        query_modes = [str(item) for item in field.get("query_modes", [])]
        if any(item not in valid_ops for item in query_modes):
            raise ValidationError(f"字段查询方式不匹配：{key}")
        enum_values = [str(item) for item in field.get("enum_values", [])]
        if field_type == "enum" and not enum_values:
            raise ValidationError(f"枚举字段必须设置选项：{key}")
        return {
            "field_key": key,
            "name": str(field.get("name") or key).strip(),
            "field_type": field_type,
            "required": bool(field.get("required", False)),
            "query_modes": query_modes,
            "sortable": bool(field.get("sortable", False)),
            "agent_readable": bool(field.get("agent_readable", True)),
            "enum_values": enum_values,
        }

    def _normalize_values(self, definition: dict[str, Any], values: dict[str, Any]) -> dict[str, Any]:
        if not isinstance(values, dict):
            raise ValidationError("记录必须是对象")
        fields = {field["field_key"]: field for field in definition["fields"]}
        unknown = sorted(set(values) - set(fields))
        if unknown:
            raise ValidationError(f"存在未定义字段：{unknown[0]}")
        result: dict[str, Any] = {}
        for key, field in fields.items():
            value = values.get(key)
            if value is None or value == "":
                if field["required"]:
                    raise ValidationError(f"必填字段不能为空：{key}")
                result[key] = None
                continue
            if field["field_type"] == "number":
                try:
                    result[key] = float(value)
                except (TypeError, ValueError) as exc:
                    raise ValidationError(f"数字字段格式无效：{key}") from exc
            elif field["field_type"] == "date":
                try:
                    result[key] = pd.Timestamp(value).date().isoformat()
                except (TypeError, ValueError) as exc:
                    raise ValidationError(f"日期字段格式无效：{key}") from exc
            elif field["field_type"] == "datetime":
                try:
                    timestamp = pd.Timestamp(value)
                    if timestamp.tzinfo is None:
                        timestamp = timestamp.tz_localize("UTC")
                    else:
                        timestamp = timestamp.tz_convert("UTC")
                    result[key] = timestamp.isoformat()
                except (TypeError, ValueError) as exc:
                    raise ValidationError(f"日期字段格式无效：{key}") from exc
            else:
                result[key] = str(value)
                if field["field_type"] == "enum" and result[key] not in field["enum_values"]:
                    raise ValidationError(f"枚举字段取值无效：{key}")
        return result

    def _list_definitions_raw(self) -> list[dict[str, Any]]:
        with self._connect() as conn:
            rows = conn.execute("SELECT * FROM business_ledgers ORDER BY ledger_key").fetchall()
        return [self._definition_row(row) for row in rows]

    def _get_definition_raw(self, ledger_key: str, *, required: bool = True) -> dict[str, Any] | None:
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM business_ledgers WHERE ledger_key = ?", (ledger_key,)).fetchone()
        if row is None:
            if required:
                raise NotFound("业务台账不存在")
            return None
        return self._definition_row(row)

    def _definition_row(self, row: sqlite3.Row) -> dict[str, Any]:
        payload = json.loads(row["definition_json"])
        return {**dict(row), **payload}

    def _records_raw(self, ledger_key: str) -> list[sqlite3.Row]:
        with self._connect() as conn:
            return conn.execute("SELECT * FROM business_ledger_records WHERE ledger_key = ? ORDER BY created_at, record_id", (ledger_key,)).fetchall()

    def _record_count(self, ledger_key: str) -> int:
        with self._connect() as conn:
            return int(conn.execute("SELECT COUNT(*) FROM business_ledger_records WHERE ledger_key = ?", (ledger_key,)).fetchone()[0])

    @staticmethod
    def _payload(definition: dict[str, Any], *, record_count: int) -> dict[str, Any]:
        return attach_edit_token({**definition, "record_count": record_count}, definition)

    @staticmethod
    def _json_value(value: Any) -> Any:
        if pd.isna(value):
            return None
        if hasattr(value, "isoformat"):
            return value.isoformat()
        return value.item() if hasattr(value, "item") else value

    @staticmethod
    def _validate_ledger_key(value: str, *, label: str = "台账标识") -> None:
        if not value or len(value) > 80 or any(char not in "abcdefghijklmnopqrstuvwxyz0123456789_-" for char in value):
            raise ValidationError(f"{label}仅支持小写字母、数字、下划线和连字符")
