from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook


def workbook_bytes(
    headers: list[object],
    rows: list[list[object]],
    *,
    second_sheet_rows: list[list[object]] | None = None,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "tasks"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    if second_sheet_rows is not None:
        ignored = workbook.create_sheet("ignored")
        for row in second_sheet_rows:
            ignored.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def malformed_worksheet_workbook_bytes() -> bytes:
    source_bytes = workbook_bytes(["task_key", "task_version", "type"], [])
    output = BytesIO()
    with ZipFile(BytesIO(source_bytes)) as source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for name in source.namelist():
            data = source.read(name)
            if name == "xl/worksheets/sheet1.xml":
                data = b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><dimension ref="A1:C1"/><sheetData>'
            target.writestr(name, data)
    return output.getvalue()


def test_parse_task_import_maps_fixed_and_extra_columns_from_first_sheet():
    from agent_bridge.automation.workflows.task_import import parse_task_import

    result = parse_task_import(
        workbook_bytes(
            ["task_key", "task_version", "type", "repo", "priority"],
            [[" owner/repo ", "v1", "repo", "owner/repo", 3]],
            second_sheet_rows=[["ignored"]],
        ),
        filename="tasks.xlsx",
    )

    assert result.sheet_name == "tasks"
    assert result.rows[0].row_number == 2
    assert result.rows[0].task_key == "owner/repo"
    assert result.rows[0].task_version == "v1"
    assert result.rows[0].task_type == "repo"
    assert result.rows[0].payload == {"repo": "owner/repo", "priority": 3}
    assert result.rows[0].errors == ()
    assert len(result.rows) == 1


def test_parse_task_import_normalizes_fixed_headers_and_scalar_values():
    from agent_bridge.automation.workflows.task_import import parse_task_import

    result = parse_task_import(
        workbook_bytes(
            ["\ufeff TASK_KEY ", " TASK_VERSION", "TYPE ", "enabled", "count", "ratio", "when"],
            [[" key ", None, None, True, 4, 2.5, date(2026, 7, 15)]],
        ),
        filename="TASKS.XLSX",
    )

    row = result.rows[0]
    assert row.task_key == "key"
    assert row.task_version == ""
    assert row.task_type == ""
    assert row.payload == {
        "enabled": True,
        "count": 4,
        "ratio": 2.5,
        "when": "2026-07-15",
    }


def test_parse_task_import_converts_datetime_and_time_values_to_isoformat():
    from agent_bridge.automation.workflows.task_import import parse_task_import

    result = parse_task_import(
        workbook_bytes(
            ["task_key", "task_version", "type", "started", "at"],
            [["task:1", "v1", "job", datetime(2026, 7, 15, 12, 34, 56), time(9, 8, 7)]],
        ),
        filename="tasks.xlsx",
    )

    assert result.rows[0].payload == {
        "started": "2026-07-15T12:34:56",
        "at": "09:08:07",
    }


def test_parse_task_import_ignores_blank_rows_and_excludes_fixed_columns_from_payload():
    from agent_bridge.automation.workflows.task_import import parse_task_import

    result = parse_task_import(
        workbook_bytes(
            ["task_key", "task_version", "type", "optional"],
            [
                [None, None, None, None],
                ["task:1", "v1", "job", None],
                ["", "", "", ""],
            ],
        ),
        filename="tasks.xlsx",
    )

    assert [row.row_number for row in result.rows] == [3]
    assert result.rows[0].payload == {}


def test_parse_task_import_reports_value_under_empty_extra_header_with_excel_row_number():
    from agent_bridge.automation.workflows.task_import import parse_task_import

    result = parse_task_import(
        workbook_bytes(
            ["task_key", "task_version", "type", None],
            [["task:1", "v1", "job", "value"]],
        ),
        filename="tasks.xlsx",
    )

    row = result.rows[0]
    assert row.row_number == 2
    assert row.payload == {}
    assert len(row.errors) == 1
    assert "第4列" in row.errors[0]


def test_parse_task_import_keeps_blank_key_row_error():
    from agent_bridge.automation.workflows.task_import import parse_task_import

    result = parse_task_import(
        workbook_bytes(
            ["task_key", "task_version", "type"],
            [[" ", "v1", "repo"]],
        ),
        filename="tasks.xlsx",
    )

    assert result.rows[0].row_number == 2
    assert result.rows[0].task_key == ""
    assert result.rows[0].errors == ("task_key 不能为空",)


def test_parse_task_import_marks_duplicate_normalized_task_key_and_version():
    from agent_bridge.automation.workflows.task_import import parse_task_import

    result = parse_task_import(
        workbook_bytes(
            ["task_key", "task_version", "type"],
            [
                [" owner/repo ", " v1 ", "repo"],
                ["owner/repo", "v1", "repo"],
            ],
        ),
        filename="tasks.xlsx",
    )

    assert result.rows[0].errors == ()
    assert result.rows[1].row_number == 3
    assert any("重复" in error for error in result.rows[1].errors)


@pytest.mark.parametrize(
    ("headers", "message"),
    [
        (["task_version", "type"], "task_key"),
        (["task_key", " TASK_KEY ", "type"], "重复"),
        (["task_key", "type", " TYPE "], "重复"),
    ],
)
def test_parse_task_import_rejects_missing_or_duplicate_normalized_headers(headers, message):
    from agent_bridge.automation.workflows.task_import import TaskImportFormatError, parse_task_import

    with pytest.raises(TaskImportFormatError, match=message):
        parse_task_import(workbook_bytes(headers, [["value", "job", "extra"]]), filename="tasks.xlsx")


def test_parse_task_import_rejects_invalid_extension_and_corrupt_zip():
    from agent_bridge.automation.workflows.task_import import TaskImportFormatError, parse_task_import

    with pytest.raises(TaskImportFormatError, match="xlsx"):
        parse_task_import(b"not an xlsx", filename="tasks.csv")

    with pytest.raises(TaskImportFormatError, match="无效"):
        parse_task_import(b"not an xlsx", filename="tasks.xlsx")


def test_parse_task_import_rejects_malformed_worksheet_xml():
    from agent_bridge.automation.workflows.task_import import TaskImportFormatError, parse_task_import

    with pytest.raises(TaskImportFormatError, match="无效"):
        parse_task_import(malformed_worksheet_workbook_bytes(), filename="tasks.xlsx")


def test_parse_task_import_enforces_maximum_of_5000_non_empty_rows():
    from agent_bridge.automation.workflows.task_import import TaskImportFormatError, parse_task_import

    rows = [[f"task:{index}", "v1", "job"] for index in range(5001)]

    with pytest.raises(TaskImportFormatError, match="5000"):
        parse_task_import(workbook_bytes(["task_key", "task_version", "type"], rows), filename="tasks.xlsx")


def test_build_task_import_template_has_fixed_columns_and_explanation_sheet():
    from agent_bridge.automation.workflows.task_import import build_task_import_template

    workbook = load_workbook(BytesIO(build_task_import_template()), read_only=True, data_only=True)

    assert workbook.sheetnames == ["tasks", "说明"]
    tasks = workbook["tasks"]
    assert next(tasks.iter_rows(values_only=True)) == ("task_key", "task_version", "type")
    assert tasks.max_column == 3

    explanation = workbook["说明"]
    explanation_text = "\n".join(str(value or "") for row in explanation.iter_rows(values_only=True) for value in row)
    assert "payload" in explanation_text
    assert "额外" in explanation_text
    workbook.close()
