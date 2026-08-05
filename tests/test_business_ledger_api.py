from __future__ import annotations

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from io import BytesIO

from agent_bridge.api.app import create_app


def test_business_ledger_management_api(wm_paths) -> None:
    app = create_app(wm_paths, {"root"})
    with TestClient(app) as client:
        headers = {"X-Agent-Bridge-User": "root"}
        created = client.post(
            "/business-ledgers",
            headers=headers,
            json={
                "ledger_key": "assets",
                "name": "资产台账",
                "fields": [
                    {"field_key": "name", "name": "名称", "field_type": "text", "query_modes": ["contains"]},
                    {"field_key": "cost", "name": "成本", "field_type": "number"},
                ],
            },
        )
        assert created.status_code == 200
        added = client.post("/business-ledgers/assets/records", headers=headers, json={"values": {"name": "支付系统", "cost": 12}})
        assert added.status_code == 200
        found = client.post(
            "/business-ledgers/assets/records/query",
            headers=headers,
            json={"keyword": "支付", "filters": {"cost": {"op": "between", "from": 10, "to": 20}}},
        )
        assert found.status_code == 200
        assert found.json()["total"] == 1

        ordered = client.post(
            "/business-ledgers/assets/records/query",
            headers=headers,
            json={"sort": [{"field": "cost", "direction": "desc"}, {"field": "name", "direction": "asc"}]},
        )
        assert ordered.status_code == 200


def test_business_ledger_excel_preview_confirm_and_export(wm_paths) -> None:
    app = create_app(wm_paths, {"root"})
    with TestClient(app) as client:
        headers = {"X-Agent-Bridge-User": "root"}
        client.post("/business-ledgers", headers=headers, json={"ledger_key": "assets", "name": "资产", "fields": [{"field_key": "name", "name": "名称", "field_type": "text", "query_modes": ["contains"]}]})
        workbook = Workbook()
        workbook.active.append(["名称"])
        workbook.active.append(["Excel 资产"])
        content = BytesIO(); workbook.save(content)
        preview = client.post("/business-ledgers/assets/imports/xlsx/preview", headers=headers, files={"file": ("assets.xlsx", content.getvalue())})
        assert preview.json()["rows"] == 1
        confirmed = client.post(f"/business-ledgers/assets/imports/xlsx/{preview.json()['preview_id']}/confirm", headers=headers)
        assert confirmed.json()["imported"] == 1
        exported = client.get("/business-ledgers/assets/exports/xlsx", headers=headers)
        assert exported.status_code == 200
        assert load_workbook(BytesIO(exported.content)).active.max_row == 2

        template = client.get("/business-ledgers/assets/imports/xlsx/template", headers=headers)
        assert template.status_code == 200
        assert template.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        template_sheet = load_workbook(BytesIO(template.content)).active
        assert template_sheet.max_row == 1
        assert [cell.value for cell in template_sheet[1]] == ["name"]
